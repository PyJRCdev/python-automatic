from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.drawing.image import Image
import time
import getpass

# Ruta a EdgeDriver
edgedriver_path = "servicios-auto/edgedriver/msedgedriver.exe"
output_excel_path = "servicios-auto/Download/niveles_servicio.xlsx"
url = 'https://portal.ismgroup.es/'
header_image_path = "servicios-auto/header_image.png"


def download_excel_with_selenium(email, password):
    # Configuración de Microsoft Edge en modo headless (sin interfaz gráfica)
    options = webdriver.EdgeOptions()
    # options.add_argument('--headless')  # Si quieres que sea sin interfaz gráfica
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')

    # Iniciar Edge
    service = Service(executable_path=edgedriver_path)
    driver = webdriver.Edge(service=service, options=options)
    driver.get(url)

    try:
        # Esperar a que cargue la página y hacer el inicio de sesión
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, 'tbUsuario')))
        
        email_input = driver.find_element(By.NAME, 'tbUsuario')
        email_input.send_keys(email)
        
        password_input = driver.find_element(By.NAME, 'tbContrasena')
        password_input.send_keys(password)
        password_input.send_keys(Keys.RETURN)

        # Esperar a que cargue la página después de iniciar sesión
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'ContentPlaceHolder1_LinkButton1')))
        
        # Navegar por los botones
        driver.find_element(By.ID, 'ContentPlaceHolder1_LinkButton1').click()
        time.sleep(2)

        # Esperar y manejar la nueva ventana
        WebDriverWait(driver, 20).until(lambda d: len(d.window_handles) > 1)
        original_window = driver.current_window_handle
        all_windows = driver.window_handles
        for window in all_windows:
            if window != original_window:
                driver.switch_to.window(window)
                break

        # Usar ActionChains para mover el ratón al botón y hacer clic en el menú desplegable
        actions = ActionChains(driver)
        popout_button = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'popout'))
        )
        actions.move_to_element(popout_button).perform()
        time.sleep(2)  # Esperar a que el menú desplegable aparezca

        # Hacer clic en la opción "Niveles de servicio"
        niveles_button = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//a[text()="Niveles de servicio"]'))
        )
        niveles_button.click()

        # Seleccionar la opción "Proyecto" del selector
        project_selector = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'ContenidoPrincipal_ContenidoServicios_ddlProyecto'))
        )
        project_selector.click()

        project_option = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//option[text()="TF Unicaja Micro"]'))
        )
        project_option.click()

        # Pulsar el botón para buscar
        buscar_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, 'ContenidoPrincipal_ContenidoServicios_bBuscar'))
        )
        buscar_button.click()
        # Pulsar el botón para imprimir/exportar
        print_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, 'ContenidoPrincipal_ContenidoServicios_bImprimir'))
        )
        print_button.click()

        # Cambiar a la nueva ventana que se abre
        WebDriverWait(driver, 20).until(lambda d: len(d.window_handles) > 1)
        all_windows = driver.window_handles
        for window in all_windows:
            if window != original_window:
                driver.switch_to.window(window)
                break
        # Esperar a que la tabla esté presente
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, 'table'))
        )
        time.sleep(20)  # Añadir tiempo para cargar completamente

        # Detectar la(s) tabla(s)
        tables = driver.find_elements(By.TAG_NAME, 'table')

        if tables:
            print(f"Se encontraron {len(tables)} tabla(s). Procesando la primera tabla.")
            table = tables[0]

            # Extraer los datos de la tabla
            rows = table.find_elements(By.TAG_NAME, 'tr')
            table_data = []
            headers = []
            for row_index, row in enumerate(rows):
                cells = row.find_elements(By.TAG_NAME, 'td' if row_index > 0 else 'th')  # Usar <th> para encabezados
                if cells:  # Evitar filas vacías
                    if row_index == 0:
                        headers = [cell.text for cell in cells]
                    else:
                        table_data.append([cell.text for cell in cells])
                        
            if table_data:
                # Crear DataFrame con los datos extraídos
                df = pd.DataFrame(table_data, columns=headers)

                # Guardar los datos en un archivo Excel
                df.to_excel(output_excel_path, index=False, header=True)
                print(f"Datos extraídos y guardados en {output_excel_path}.")

                # Cargar el archivo Excel para aplicar formato
                wb = load_workbook(output_excel_path)
                ws = wb.active

                # Insertar una fila vacía para la imagen
                ws.insert_rows(1)

                # Eliminar columnas vacías
                for col in ws.columns:
                    if all(cell.value is None for cell in col):
                        ws.delete_cols(col[0].column)
                
                # Ajustar altura de la fila para la imagen
                ws.row_dimensions[1].height = 55

                # Eliminar columnas vacías
                max_col = ws.max_column
                max_row = ws.max_row

                # Iterar sobre cada columna de derecha a izquierda para evitar problemas al eliminar
                for col in range(max_col, 0, -1):
                    if all(ws.cell(row=row, column=col).value is None for row in range(2, max_row + 1)):
                        ws.delete_cols(col)

                N = 2
                # Aplicar formato a la primera fila
                fill = PatternFill(start_color='267cb2', end_color='267cb2', fill_type='solid')
                font = Font(color='FFFFFF')
                for row in range(1, N + 1):  # Comienza en la fila 2 porque la fila 1 tiene la imagen
                    for cell in ws[row]:
                        cell.fill = fill
                        cell.font = font

                # Añadir una imagen como encabezado
                img = Image(header_image_path)
                ws.add_image(img, 'A1')

                # Guardar el archivo Excel con el formato aplicado
                wb.save(output_excel_path)
                print(f"Formato aplicado y archivo guardado en {output_excel_path}.")
            else:
                print("No se encontraron datos en la tabla.")
        else:
            print("No se encontró ninguna tabla en la página.")


    finally:
        driver.quit()


email = input("Introduce tu email: ")
password = getpass.getpass("Introduce tu contraseña: ")

# Ejecutar el proceso
download_excel_with_selenium(email, password)

print("El proceso ha finalizado.")
